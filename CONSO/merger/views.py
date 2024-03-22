import os
from django.shortcuts import render
from .forms import MergeRequestForm
from .models import MergeRequest
from openpyxl import load_workbook
import mimetypes
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

def home(request):
    return render(request, 'merger/index.html')

def preview_file(request, pk, field):
    merge_request = get_object_or_404(MergeRequest, pk=pk)
    file_path = getattr(merge_request, field).path
    file_name = os.path.basename(file_path)
    file_type, _ = mimetypes.guess_type(file_path)

    with open(file_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type=file_type)
        response['Content-Disposition'] = f'inline; filename="{file_name}"'
        return response

def merge_excel(request):
    if request.method == 'POST':
        form = MergeRequestForm(request.POST, request.FILES)
        if form.is_valid():
            merge_request = form.save()
            file1 = merge_request.file1.path
            file2 = merge_request.file2.path

            workbook1 = load_workbook(filename=file1)
            workbook2 = load_workbook(filename=file2)

            worksheet1 = workbook1.active
            worksheet2 = workbook2.active

            merged_workbook = load_workbook(filename=file1)
            merged_worksheet = merged_workbook.active

            for row in worksheet2.iter_rows():
                merged_worksheet.append([cell.value for cell in row])

            output_path = os.path.join(os.path.dirname(file1), 'merged.xlsx')
            merged_workbook.save(output_path)

            os.remove(file1)
            os.remove(file2)

            context = {
                'form': form,
                'output_path': output_path,
            }
            return render(request, 'merger/merge_success.html', context)
    else:
        form = MergeRequestForm()

    return render(request, 'merger/merge_form.html', {'form': form})